from requests import get
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
 
input_file_name = input('Enter name of Excel file to read (include .xlsx or other extension): ')

file_data = openpyxl.load_workbook(input_file_name)
account_data = file_data.active

max_row = account_data.max_row

# Data arrays
names = []
usernames = []
profile_urls = []
followers_counts = []

# Get all names
for row in account_data.iter_rows(min_row=2, min_col=1, max_row=max_row, max_col=1):
  for cell in row:
    names.append(cell.value)

# Get all usernames
for row in account_data.iter_rows(min_row=2, min_col=2, max_row=max_row, max_col=2):
  for cell in row:
    profile_urls.append(cell.value)
    username = cell.value.split("/")[3]
    usernames.append(username)

# Excel columns
col1 = 'Name'
col2 = 'Username'
col3 = 'Profile URL'
col4 = 'Follower Count'

for username in usernames:
  try:
    url = '%s/%s/' % ('https://www.instagram.com', username)
    page = get(url, timeout=5)
    soup = BeautifulSoup(page.content, 'html.parser')
    data = soup.find_all('meta', attrs={'property': 'og:description'})
    text = data[0].get('content').split()
    follower_count = text[0]
    followers_counts.append(follower_count)
  except:
    print('Error scraping data for user: ' + username)

data = pd.DataFrame({col1: names, col2: usernames, col3: profile_urls, col4: followers_counts})

file_name = input('Enter a file name for the Excel file that will contain your scraped data (do not include the .xlsx suffix): ')

sheet_name = input('Enter a name for the Excel sheet: ')

data.to_excel(f'{file_name}.xlsx', sheet_name=sheet_name)

input('The .xlsx file has been created and saved in the projects root directory.')
