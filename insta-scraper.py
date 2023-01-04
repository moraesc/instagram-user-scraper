import instaloader
import pandas as pd
import openpyxl
import re

bot = instaloader.Instaloader()
id = 'ADD_YOUR_USERNAME_HERE'
pw = 'ADD_YOUR_PASSWORD_HERE!'
bot.login(id, pw)

inputfileName = input('Enter name of Excel file to read (include .xlsx or other extension): ')

data = openpyxl.load_workbook(inputfileName)
accountData = data.active

maxRow = accountData.max_row

# Data arrays
usernames = []
followers = []
followees = []
posts = []

for row in accountData.iter_rows(min_row=2, min_col=2, max_row=maxRow, max_col=2):
  for cell in row:
    if cell.value is not None:
      username = cell.value.split("/")[3]
      usernames.append(username)

# Excel columns
col1 = 'Username'
col2 = 'Followers'
col3 = 'Followees'
col4 = 'Number of Posts'

for username in usernames:
  try:
    profile = instaloader.Profile.from_username(bot.context, username)
    usernames.append(profile.username)
    followers.append(profile.followers)
    followees.append(profile.followees)
    posts.append(profile.mediacount)
  except:
    print('Profile for username ' + username + ' does not exist')

data = pd.DataFrame({col1: usernames, col2: followers, col3: followees, col4: posts})

fileName = input('Enter a file name for the Excel file that will contain your scraped data (do not include the .xlsx suffix): ')

sheetName = input('Enter a name for the Excel sheet: ')

data.to_excel(f'{fileName}.xlsx', sheet_name=sheetName)

